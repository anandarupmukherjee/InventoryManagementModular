import io
import csv
import zipfile
import datetime
from django.apps import apps
from django.http import HttpResponse
from django.shortcuts import render
from django.utils.timezone import make_aware
from openpyxl import Workbook
from openpyxl.styles import Font

from services.data_storage.models import (
    Product,
    Withdrawal,
    PurchaseOrder,
    StockRegistrationLog,
)  # Adjust as needed


MODEL_DEFINITIONS = {
    'Withdrawal': {
        'label': 'Withdrawals',
        'model_path': 'data_storage.Withdrawal',
        'date_field': 'timestamp',
    },
    'Product': {
        'label': 'Products',
        'model_path': 'data_storage.Product',
    },
    'PurchaseOrder': {
        'label': 'Purchase Orders',
        'model_path': 'data_storage.PurchaseOrder',
        'date_field': 'order_date',
    },
    'StockRegistrationLog': {
        'label': 'Stock Registration Logs (Check-ins)',
        'model_path': 'data_storage.StockRegistrationLog',
        'date_field': 'timestamp',
    },
    'DeletedLot': {
        'label': 'Deleted Lots',
        'model_path': 'data_storage.Withdrawal',
        'date_field': 'timestamp',
        'base_filters': {
            'withdrawal_type': 'lot_discard',
        },
    },
}


MODEL_MAP = {name: meta['model_path'] for name, meta in MODEL_DEFINITIONS.items()}
MODEL_LABELS = {name: meta['label'] for name, meta in MODEL_DEFINITIONS.items()}
MODEL_CHOICES = [(name, meta['label']) for name, meta in MODEL_DEFINITIONS.items()]


FILTER_FIELDS = {
    name: meta['date_field']
    for name, meta in MODEL_DEFINITIONS.items()
    if meta.get('date_field')
}

def download_report(request):
    default_model = MODEL_CHOICES[0][0]
    selected_model = request.GET.get('model', default_model)
    if selected_model not in MODEL_MAP:
        selected_model = default_model
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    download_type = request.GET.get('download')

    preview_model_class = apps.get_model(*MODEL_MAP[selected_model].split('.'))
    base_filters = MODEL_DEFINITIONS[selected_model].get('base_filters', {})
    preview_fields = [f.name for f in preview_model_class._meta.fields]

    if base_filters:
        preview_queryset = preview_model_class.objects.filter(**base_filters)
    else:
        preview_queryset = preview_model_class.objects.all()

    # Filter preview table if model supports date filtering
    if selected_model in FILTER_FIELDS:
        date_field = FILTER_FIELDS[selected_model]
        if start_date:
            preview_queryset = preview_queryset.filter(
                **{f"{date_field}__gte": make_aware(datetime.datetime.strptime(start_date, "%Y-%m-%d"))}
            )
        if end_date:
            preview_queryset = preview_queryset.filter(
                **{f"{date_field}__lte": make_aware(datetime.datetime.strptime(end_date, "%Y-%m-%d"))}
            )

    # ✅ THEN slice it
    preview_queryset = preview_queryset.order_by('-id')


    # Handle Excel or CSV download (only for the selected model + date range)
    if download_type in ['excel', 'csv']:
        model_path = MODEL_MAP[selected_model]
        model_class = apps.get_model(*model_path.split('.'))
        fields = [f.name for f in model_class._meta.fields]
        qs = model_class.objects.filter(**base_filters) if base_filters else model_class.objects.all()

        if selected_model in FILTER_FIELDS:
            date_field = FILTER_FIELDS[selected_model]
            if start_date:
                qs = qs.filter(
                    **{f"{date_field}__gte": make_aware(datetime.datetime.strptime(start_date, "%Y-%m-%d"))}
                )
            if end_date:
                qs = qs.filter(
                    **{f"{date_field}__lte": make_aware(datetime.datetime.strptime(end_date, "%Y-%m-%d"))}
                )

        # Prepare meta
        user_name = getattr(request.user, 'username', '') or 'Anonymous'
        today_str = datetime.date.today().strftime('%Y-%m-%d')
        if start_date or end_date:
            range_str = f"{start_date or '—'} to {end_date or '—'}"
        else:
            range_str = "All dates"

        if download_type == 'excel':
            wb = Workbook()
            # openpyxl creates a default sheet; write into it and rename for cleanliness
            ws = wb.active
            ws.title = selected_model
            # Top meta rows (bold)
            bold = Font(bold=True)
            ws.append(["Downloaded by", user_name])
            ws.append(["Download date", today_str])
            ws.append(["Date range", range_str])
            for cell in ws[1] + ws[2] + ws[3]:
                cell.font = bold
            ws.append([])  # blank spacer
            # Header row
            ws.append(fields)
            for cell in ws[5]:
                cell.font = bold
            for obj in qs:
                ws.append([str(getattr(obj, field, '')) for field in fields])

            out = io.BytesIO()
            wb.save(out)
            out.seek(0)
            response = HttpResponse(out, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = f'attachment; filename=report_{selected_model.lower()}_{today_str}.xlsx'
            return response

        elif download_type == 'csv':
            csv_buffer = io.StringIO()
            writer = csv.writer(csv_buffer)
            # Meta rows
            writer.writerow(["Downloaded by", user_name])
            writer.writerow(["Download date", today_str])
            writer.writerow(["Date range", range_str])
            writer.writerow([])
            # Header row
            writer.writerow(fields)
            for obj in qs:
                writer.writerow([str(getattr(obj, field, '')) for field in fields])
            response = HttpResponse(csv_buffer.getvalue(), content_type='text/csv')
            response['Content-Disposition'] = f'attachment; filename=report_{selected_model.lower()}_{today_str}.csv'
            return response

    return render(request, 'inventory/download_report.html', {
        'model_options': MODEL_CHOICES,
        'selected_model': selected_model,
        'selected_model_label': MODEL_LABELS.get(selected_model, selected_model),
        'fields': preview_fields,
        'data': preview_queryset,
        'start_date': start_date,
        'end_date': end_date,
    })
